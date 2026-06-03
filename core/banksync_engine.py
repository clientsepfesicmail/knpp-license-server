"""
BankSync Pro v2 — Single File Application
PDF Bank Statement → Tally Prime
"""

import threading
import os
import json
import re
import pdfplumber
try:
    import pikepdf
    PIKEPDF_AVAILABLE = True
except ImportError:
    PIKEPDF_AVAILABLE = False
try:
    import pandas as pd
except ImportError:  # Phase 3 preview does not require CSV export.
    pd = None
import xml.etree.ElementTree as ET
from xml.dom import minidom
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:  # Phase 3 preview does not require Excel export.
    openpyxl = None
    Font = PatternFill = Alignment = Border = Side = None
from datetime import datetime


def _get_user_data_dir():
    base = os.environ.get("APPDATA") or os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    path = os.path.join(base, "BankImportPro")
    os.makedirs(path, exist_ok=True)
    return path

PROFILES_FILE = os.path.join(_get_user_data_dir(), "profiles.json")

def _migrate_profiles_file():
    try:
        old = os.path.join(os.getcwd(), "profiles.json")
        if os.path.exists(old) and not os.path.exists(PROFILES_FILE):
            import shutil
            shutil.copy2(old, PROFILES_FILE)
    except Exception:
        pass

_migrate_profiles_file()


# ═══════════════════════════════════════════════════════════════
#  CORE ENGINE
# ═══════════════════════════════════════════════════════════════

# ── PDF UNLOCK ───────────────────────────────────────────────
def unlock_pdf(pdf_path, password=""):
    """Password-protected PDF unlock karo. pikepdf optional hai."""
    if not password:
        return pdf_path  # No password — direct use karo
    if not PIKEPDF_AVAILABLE:
        # pikepdf nahi hai — as-is return karo, pdfplumber try karega
        return pdf_path
    unlocked_path = pdf_path.replace(".pdf", "_unlocked.pdf")
    try:
        with pikepdf.open(pdf_path, password=password) as pdf:
            pdf.save(unlocked_path)
        return unlocked_path
    except Exception:
        return pdf_path


# ── PDF EXTRACT ──────────────────────────────────────────────
def extract_pdf(pdf_path, progress_cb=None, password=""):
    all_text, all_tables = [], []
    open_kwargs = {"password": password} if password else {}
    try:
        with pdfplumber.open(pdf_path, **open_kwargs) as pdf:
            total = len(pdf.pages)
            for i, page in enumerate(pdf.pages):
                if progress_cb:
                    progress_cb(i + 1, total)
                text = page.extract_text()
                if text:
                    all_text.append(text)
                for tbl in (page.extract_tables() or []):
                    if tbl:
                        all_tables.extend(tbl)
    except Exception as e:
        msg = str(e).strip()
        err_text = f"{type(e).__name__} {msg}".lower()
        if any(k in err_text for k in ["password", "encrypted", "decrypt"]):
            raise RuntimeError("This PDF appears to be password-protected. Please enter the PDF password and try again.") from e
        if not msg:
            raise RuntimeError(f"Could not read the PDF ({type(e).__name__}). If the statement is password-protected, enter the password and retry.") from e
        raise RuntimeError(f"Could not read the PDF: {type(e).__name__}: {msg}") from e
    return "\n".join(all_text), all_tables



def extract_pdf_text_fast(pdf_path, progress_cb=None, password=""):
    """
    Fast text-only extractor for long text-based statements. HDFC PDFs often
    parse better from raw text because table extraction may collapse
    Withdrawal/Deposit columns into one amount column.
    """
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        if getattr(reader, "is_encrypted", False):
            if not password:
                return ""
            try:
                reader.decrypt(password)
            except Exception:
                return ""
        total = len(reader.pages)
        parts = []
        for i, page in enumerate(reader.pages):
            if progress_cb:
                progress_cb(i + 1, total)
            try:
                page_text = page.extract_text() or ""
            except Exception:
                page_text = ""
            if page_text:
                parts.append(page_text)
        return "\n".join(parts).strip()
    except Exception:
        return ""



# ── BANK DETECT ──────────────────────────────────────────────
def detect_bank(text):
    """
    Auto bank detection ko sirf transaction narration pe depend mat karo.
    Axis jaise statements me transaction details ke andar doosre banks ke naam aate hain,
    isliye header / IFSC / scheme based weighted detection use karo.
    """
    raw = str(text or "")
    t = raw.lower()
    header = raw[:2500].lower()

    # Strong clues from the true statement header / IFSC / MICR section.
    # Use only the top part first because transaction narration may mention
    # many other banks (Axis statements can contain UCO/ICICI/HDFC in UPI rows).
    top_header = raw[:1000].lower()

    if "south indian bank" in top_header or re.search(r"ifsc\s*:?\s*sibl", top_header) or re.search(r"\bsibl\d{4}", top_header):
        return "SOUTH INDIAN BANK"
    if re.search(r"ifsc\s*code\s*:?\s*utib", top_header) or re.search(r"micr/ifsc\s+code\s*:?\s*[\d\s/]+utib", top_header) or re.search(r"\butib\d{4}", top_header):
        return "AXIS"
    if re.search(r"ifsc\s*code\s*:?\s*idib", top_header) or re.search(r"\bidib\d{4}\b", top_header):
        return "INDIAN BANK"
    if re.search(r"ifsc\s*code\s*:?\s*sbin", top_header) or re.search(r"\bsbin\d{4}\b", top_header):
        return "SBI"
    if re.search(r"ifsc\s*code\s*:?\s*hdfc", top_header) or re.search(r"\bhdfc\d{4}\b", top_header):
        return "HDFC"
    if re.search(r"ifsc\s*code\s*:?\s*icic", top_header) or re.search(r"\bicic\d{4}\b", top_header):
        return "ICICI"
    if re.search(r"ifsc\s*code\s*:?\s*ibkl", top_header) or re.search(r"\bibkl\d{4}\b", top_header) or "idbi bank" in top_header:
        return "IDBI"
    if re.search(r"ifsc\s*code\s*:?\s*kkbk", top_header) or re.search(r"\bkkbk\d{4}\b", top_header):
        return "KOTAK"
    if re.search(r"ifsc\s*code\s*:?\s*punb", top_header) or re.search(r"\bpunb\d{4}\b", top_header):
        return "PNB"
    if re.search(r"ifsc\s*code\s*:?\s*barb", top_header) or re.search(r"\bbarb\d{4}\b", top_header):
        return "BOB"
    if "uco bank" in top_header or re.search(r"ifsc\s*code\s*:?\s*ucba", top_header) or re.search(r"\bucba\d{4}\b", top_header):
        return "UCO"

    weighted_banks = {
        "UCO": {
            "header": ["uco bank", "uco bank account name", "account name :"],
            "body": ["uco bank"],
        },
        "IDBI": {
            "header": ["idbi bank", "statement of account :", "branch ifsc code : ibkl", "branch ifsc code :ibkl", "statement generated by :"],
            "body": ["idbi bank", "ibkl"],
        },
        "SOUTH INDIAN BANK": {
            "header": ["south indian bank", "south indian bank ltd", "ifsc : sibl", "ifsc:sibl", "statement of account for the period"],
            "body": ["south indian bank", "sibl"],
        },
        "INDIAN BANK": {
            "header": ["indian bank", "branch code :", "product:", "drawing power:", "statement date :"],
            "body": [" indian bank ", "indianbank.co.in", "idib"],
        },
        "SBI": {
            "header": ["state bank of india", " ifsc code :sbin", " ifsc code : sbin"],
            "body": ["state bank of india", "sbi"],
        },
        "HDFC": {
            "header": ["hdfc bank", " ifsc code :hdfc", " ifsc code : hdfc"],
            "body": ["hdfc bank"],
        },
        "ICICI": {
            "header": ["icici bank", " ifsc code :icic", " ifsc code : icic"],
            "body": ["icici bank"],
        },
        "AXIS": {
            "header": ["axis bank", "scheme :ca - business classic", "statement of account no :", "init.\nbr", " ifsc code :utib", " ifsc code : utib"],
            "body": ["axis bank"],
        },
        "KOTAK": {
            "header": ["kotak mahindra", "kotak bank", " ifsc code :kkbk", " ifsc code : kkbk"],
            "body": ["kotak mahindra", "kotak bank"],
        },
        "PNB": {
            "header": ["punjab national", " ifsc code :punb", " ifsc code : punb"],
            "body": ["punjab national", " pnb "],
        },
        "BOB": {
            "header": ["bank of baroda", " ifsc code :barb", " ifsc code : barb"],
            "body": ["bank of baroda", " bank of baroda ", " bob "],
        },
    }

    scores = {}
    for name, clues in weighted_banks.items():
        score = 0
        for clue in clues.get("header", []):
            if clue in header:
                score += 5
        for clue in clues.get("body", []):
            if clue in t:
                score += 1
        scores[name] = score

    best_bank = max(scores, key=scores.get) if scores else "UNKNOWN"
    return best_bank if scores.get(best_bank, 0) > 0 else "UNKNOWN"


# ── DATE PARSE ───────────────────────────────────────────────
def parse_date(s):
    s = str(s).strip()
    s = re.sub(r"(st|nd|rd|th)\b", "", s, flags=re.IGNORECASE).strip()
    for fmt in [
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
        "%d/%m/%y", "%d-%m-%y", "%d.%m.%y",
        "%Y-%m-%d",
        "%d %b %Y", "%d %B %Y", "%d-%b-%Y", "%d-%B-%Y", "%d/%b/%Y", "%d/%B/%Y",
        "%d %b %y", "%d %B %y", "%d-%b-%y", "%d-%B-%y", "%d/%b/%y", "%d/%B/%y",
    ]:
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def format_tally_date(dt):
    return dt.strftime("%Y%m%d") if dt else "20250101"


def clean_amount(v):
    try:
        return float(str(v).replace(",", "").replace("₹", "").replace(" ", "").strip())
    except:
        return 0.0


# ── VOUCHER TYPE DETECT ──────────────────────────────────────
def detect_voucher_type(narration, debit, credit):
    return "Payment" if debit > 0 else "Receipt"


# ── HEADER DETECT ────────────────────────────────────────────
def detect_columns(tables):
    """
    Table header row se column map detect karo.
    Agar header nahi mila toh SBI default format use karo.
    SBI format: [ValueDate, PostDate, Narration, RefNo, Debit, Credit, Balance]
    """
    header_keywords = {
        "date":      ["value date", "txn date", "transaction date", "post date", "date"],
        "narration": ["narration", "description", "details", "particulars", "remarks"],
        "ref":       ["ref no", "chq no", "cheque no", "ref no/", "reference", "chq", "cheque", "ref"],
        "debit":     ["debit", "withdrawal", "withdraw"],
        "credit":    ["credit", "deposit"],
        "balance":   ["balance", "bal"],
    }
    col_map  = {}
    hdr_rows = set()

    for row_idx, row in enumerate(tables[:10]):
        if not row:
            continue
        row_lower = [str(c).lower().strip() if c else "" for c in row]
        matched   = {}
        for field, keywords in header_keywords.items():
            for i, cell in enumerate(row_lower):
                if any(kw in cell for kw in keywords) and field not in matched:
                    matched[field] = i
        if "date" in matched and ("debit" in matched or "credit" in matched):
            col_map  = matched
            hdr_rows.add(row_idx)
            break

    return col_map, hdr_rows


def _sbi_col_map():
    """SBI default: ValueDate=0, PostDate=1, Narration=2, Ref=3, Debit=4, Credit=5, Balance=6"""
    return {"date": 0, "narration": 2, "ref": 3, "debit": 4, "credit": 5, "balance": 6}


def _parse_amount(val):
    """'-' ya blank = 0, warna float"""
    v = str(val).strip()
    if v in ["-", "--", "", "NA", "N/A"]:
        return 0.0
    return clean_amount(v)


# ── PARSE FROM TABLE ─────────────────────────────────────────
def parse_from_table(tables):
    txns    = []
    col_map, hdr_rows = detect_columns(tables)

    # Header nahi mila? — SBI default try karo
    # SBI pages pe pehli row hoti hai ['', '', '', '', '', '', 'Balance']
    # Toh col_map empty rahega, hum SBI default use karenge
    use_sbi_default = not col_map

    if use_sbi_default:
        col_map = _sbi_col_map()

    for row_idx, row in enumerate(tables):
        if not row or len(row) < 3:
            continue

        row = [str(c).strip() if c else "" for c in row]

        # Skip header / balance-only rows
        row_joined = " ".join(row).lower()
        if row_idx in hdr_rows:
            continue
        if all(c in ["", "-", "balance", "bal"] for c in row_joined.split()):
            continue
        # Row jisme sirf ek cell mein "Balance" ho (SBI page header)
        non_empty = [c for c in row if c.strip()]
        if len(non_empty) == 1 and non_empty[0].lower() in ["balance", "bal"]:
            continue

        # ── Date ──
        date_obj = None
        date_col = col_map.get("date", 0)
        if date_col < len(row):
            date_obj = parse_date(row[date_col])
        if not date_obj:
            # fallback: pehla valid date cell
            for cell in row:
                date_obj = parse_date(cell)
                if date_obj:
                    break
        if not date_obj:
            continue

        # ── Narration ──
        narr_col  = col_map.get("narration", 2)
        narration = row[narr_col].strip() if narr_col < len(row) else ""
        # Newlines replace karo space se
        narration = re.sub(r"\s+", " ", narration).strip()
        # Agar narration ke start mein date aa gayi ho toh remove karo
        narration = re.sub(r"^\d{2}[\/\-]\d{2}[\/\-]\d{2,4}\s*", "", narration).strip()

        # ── Ref No — narration ke end mein ──
        ref_col = col_map.get("ref", 3)
        if ref_col < len(row):
            ref_val = row[ref_col].strip()
            if ref_val and ref_val not in ["-", "--", "NA", "N/A", ""]:
                narration = f"{narration} | Ref: {ref_val}" if narration else f"Ref: {ref_val}"

        if not narration:
            narration = "Bank Transaction"

        # ── Amounts — ONLY from dedicated columns ──
        debit_col  = col_map.get("debit",   4)
        credit_col = col_map.get("credit",  5)
        bal_col    = col_map.get("balance", 6)

        debit   = _parse_amount(row[debit_col])  if debit_col  < len(row) else 0.0
        credit  = _parse_amount(row[credit_col]) if credit_col < len(row) else 0.0
        balance = _parse_amount(row[bal_col])    if bal_col    < len(row) else 0.0

        if debit == 0 and credit == 0:
            continue

        txns.append({
            "date":         date_obj,
            "narration":    narration,
            "voucher_type": detect_voucher_type(narration, debit, credit),
            "debit":        debit,
            "credit":       credit,
            "balance":      balance,
        })


    return txns


# ── UCO BANK TABLE PARSER ────────────────────────────────────
def _uco_parse_amount(val):
    """UCO amount cell ko safe float me convert karo. Blank/None = 0."""
    v = str(val or "").strip()
    if not v or v in ("-", "--", "NA", "N/A"):
        return 0.0
    v = re.sub(r"\s+", "", v)
    v = v.replace(",", "").replace("₹", "")
    v = re.sub(r"(?i)(CR|DR)$", "", v)
    try:
        return round(float(v), 2)
    except Exception:
        return 0.0


def _uco_parse_balance(val):
    """UCO balance: '-29532702.22 DR' / '123.45 CR' ko signed balance me convert karo."""
    s = str(val or "").strip().replace(",", "")
    m = re.search(r"(-?\d+(?:\.\d{1,2})?)\s*(CR|DR)?", s, re.IGNORECASE)
    if not m:
        return 0.0
    amount = float(m.group(1))
    tag = (m.group(2) or "").upper()
    # UCO often prints overdraft/current account balance as DR. Keep DR negative.
    if tag == "DR" and amount > 0:
        amount = -amount
    return round(amount, 2)


def parse_uco_table(tables):
    """
    UCO Bank statement layout:
    [Date, Particulars, Chq/Ref No, Debit, Credit, Balance]

    Generic parser galat result de raha tha kyunki header missing hone par SBI default
    columns use ho rahe the. UCO me col 3 = debit, col 4 = credit, col 5 = balance.
    """
    txns = []
    for row in tables:
        if not row or len(row) < 6:
            continue

        cells = [str(c).strip() if c is not None else "" for c in row]
        first = cells[0].strip()
        if first.upper().startswith("GRAND TOTAL"):
            continue

        dt = parse_date(first)
        if not dt:
            continue

        narration = re.sub(r"\s+", " ", cells[1]).strip()
        ref = re.sub(r"\s+", " ", cells[2]).strip()
        if ref and ref not in ("-", "--", "NA", "N/A"):
            narration = f"{narration} | Ref: {ref}" if narration else f"Ref: {ref}"
        if not narration:
            narration = "Bank Transaction"

        debit = _uco_parse_amount(cells[3])
        credit = _uco_parse_amount(cells[4])
        balance = _uco_parse_balance(cells[5])

        if debit == 0.0 and credit == 0.0:
            continue

        txns.append({
            "date": dt,
            "narration": narration,
            "voucher_type": detect_voucher_type(narration, debit, credit),
            "debit": round(float(debit or 0), 2),
            "credit": round(float(credit or 0), 2),
            "balance": round(float(balance or 0), 2),
        })
    return txns


# ── INDIAN BANK TEXT PARSER HELPERS ───────────────────────────
def _signed_balance_from_token(token):
    s = str(token).strip()
    m = re.match(r"^([\d,]+\.\d{2})(Dr|Cr)$", s, re.IGNORECASE)
    if not m:
        return None
    value = clean_amount(m.group(1))
    sign  = 1 if m.group(2).lower() == "dr" else -1
    return round(sign * value, 2)


def _is_indian_bank_anchor_line(line):
    return bool(re.match(r"^\d{2}/\d{2}/\d{2}\s+\d{2}/\d{2}/\d{2}\b", line.strip()))


def _is_indian_bank_control_line(line):
    s = re.sub(r"\s+", " ", str(line).strip())
    u = s.upper()
    if not s:
        return True
    if _is_indian_bank_anchor_line(s):
        return False
    # Divider rows printed by the statement must never enter narration.
    if re.fullmatch(r"[-_*\s]+", s):
        return True
    prefixes = (
        "STATEMENT OF ACCOUNT", "DAS COMPANY", "ACCOUNT NO", "STATEMENT FROM", "TO :",
        "POST VALUE", "DETAILS CHQ.NO. DEBIT CREDIT BALANCE", "DATE DATE",
        "STATEMENT", "SUMMARY", "IN CASE YOUR ACCOUNT", "THE TRANSACTION WITH EXTRA CARE",
        "*** END OF STATEMENT ***", "*--- END OF STATEMENT ---*", "PAGE NO.", "PRODUCT:", "CURRENCY:", "INT RATE :",
        "LIMIT :", "DRAWING POWER:", "CLEARED BALANCE :", "UNCLEARED AMOUNT :", "SWEEP BAL :",
        "CKYC ID :", "NOMINATION REGD", "REGISTRATION NUMBER", "NAME OF THE NOMINEE",
        "NOMINEE TYPE :", "NOMINEE NAME", "OPPOSITE LICI OFFICE", "MR. ",
        "PO DHEMAJI", "DIST DHEMAJI", "DHEMAJI BRANCH", "CHARIALI,", "PHONE NO :",
        "EMAIL ID :", "IFSC CODE :", "BRANCH CODE :", "STATEMENT DATE :", "STATEMENT TIME :",
        "CR. COUNT:", "DR. COUNT:", "BROUGHT FORWARD", "CARRIED FORWARD",
    )
    if u.startswith(prefixes):
        return True
    if u in {"DHEMAJI", "787057"}:
        return True
    if re.match(r"^DR\. COUNT:\d+ CR\. COUNT:\d+", u):
        return True
    if re.match(r"^CR\. COUNT:\d+", u):
        return True
    if re.match(r"^SUMMARY COUNT:\d+", u):
        return True
    return False


def _clean_indian_bank_narration(parts):
    cleaned = []
    for raw in parts:
        s = re.sub(r"\s+", " ", str(raw).strip())
        if not s or _is_indian_bank_control_line(s):
            continue
        cleaned.append(s)
    narration = " ".join(cleaned)
    narration = re.sub(r"\s+", " ", narration).strip(" -|/")
    return narration or "Bank Transaction"


def _indian_bank_amount_side(narration, amount, prev_balance_signed, current_balance_signed):
    narration_l = narration.lower()
    debit = credit = 0.0
    if prev_balance_signed is not None and current_balance_signed is not None:
        delta = round(current_balance_signed - prev_balance_signed, 2)
        if abs(abs(delta) - round(amount, 2)) <= 2.0:
            if delta > 0:
                debit = amount
            elif delta < 0:
                credit = amount
    if debit == 0.0 and credit == 0.0:
        credit_hints = [
            "cash dep", "deposit by self", "deposit", "transfer from", "payment from",
            "chq dep", "cheque dep", "cheque deposit", "salary", "received", "upi"
        ]
        debit_hints = [
            "txn amt.", "transfer to", "charges", "debit interest", "interest", "neft/",
            "rtgs/", "imps/", "clear", "withdraw", "inspection charges"
        ]
        if any(k in narration_l for k in credit_hints):
            credit = amount
        elif any(k in narration_l for k in debit_hints):
            debit = amount
        else:
            debit = amount
    return round(debit, 2), round(credit, 2)


def parse_indian_bank_text(text):
    """
    Parse Indian Bank / erstwhile Allahabad Bank text statements.

    Important layout rule:
    Every non-control line printed after a dated amount/balance row belongs to
    that SAME transaction until the next dated amount/balance row starts.
    Earlier code shifted continuation lines such as ``CASH DEP/DHEMAJI`` and
    ``Txn Amt.`` to the following voucher, causing narration misalignment in
    Tally.  Keep them attached to their own anchor row.
    """
    lines = [re.sub(r"\s+", " ", line).strip() for line in str(text).split("\n")]
    lines = [line for line in lines if line]

    anchor_re = re.compile(
        r"^(?P<post>\d{2}/\d{2}/\d{2})\s+(?P<value>\d{2}/\d{2}/\d{2})(?:\s+(?P<middle>.*?))?\s+(?P<amount>[\d,]+\.\d{2})\s+(?P<balance>[\d,]+\.\d{2})(?P<baltype>Dr|Cr)\s*$",
        re.IGNORECASE,
    )
    summary_bal_re = re.compile(r"(?:BROUGHT FORWARD|CARRIED FORWARD|CLOSING BALANCE)\s*:?\s*([\d,]+\.\d{2})(Dr|Cr)", re.IGNORECASE)

    txns = []
    prev_balance_signed = None
    current = None
    current_extra = []

    def flush_current():
        nonlocal current, current_extra, prev_balance_signed
        if not current:
            return
        parts = []
        inline = current.get("inline", "")
        if inline:
            parts.append(inline)
        parts.extend(current_extra)
        narration = _clean_indian_bank_narration(parts)
        debit, credit = _indian_bank_amount_side(
            narration=narration,
            amount=current["amount"],
            prev_balance_signed=prev_balance_signed,
            current_balance_signed=current["balance_signed"],
        )
        txns.append({
            "date": current["date"],
            "narration": narration,
            "voucher_type": detect_voucher_type(narration, debit, credit),
            "debit": debit,
            "credit": credit,
            "balance": abs(current["balance_signed"]),
        })
        prev_balance_signed = current["balance_signed"]
        current = None
        current_extra = []

    for line in lines:
        bal_match = summary_bal_re.search(line)
        if bal_match and not _is_indian_bank_anchor_line(line):
            flush_current()
            prev_balance_signed = _signed_balance_from_token(f"{bal_match.group(1)}{bal_match.group(2)}")
            continue

        anchor_match = anchor_re.match(line)
        if anchor_match:
            flush_current()
            current = {
                "date": parse_date(anchor_match.group("value")),
                "inline": (anchor_match.group("middle") or "").strip(),
                "amount": clean_amount(anchor_match.group("amount")),
                "balance_signed": _signed_balance_from_token(f"{anchor_match.group('balance')}{anchor_match.group('baltype')}") or 0.0,
            }
            current_extra = []
            continue

        if _is_indian_bank_control_line(line):
            continue

        # Continuation rows are printed immediately below their own dated row.
        # Ignore pre-transaction text, but keep all useful lines after an anchor.
        if current is not None:
            current_extra.append(line)

    flush_current()
    return txns


# ── SOUTH INDIAN BANK TEXT PARSER ────────────────────────────
def _sib_signed_balance(token):
    s = str(token or "").strip()
    m = re.search(r"([\d,]+\.\d{2})\s*(Cr|Dr)\b", s, re.IGNORECASE)
    if not m:
        return None
    value = clean_amount(m.group(1))
    # Bank statement balance is positive for Cr and negative for Dr.
    return round(value if m.group(2).lower() == "cr" else -value, 2)


def _is_sib_control_line(line):
    s = re.sub(r"\s+", " ", str(line or "").strip())
    if not s:
        return True
    u = s.upper()
    if set(s) <= {"-"}:
        return True

    # South Indian Bank Finacle statements print footer/summary lines after the
    # last transaction on the same page.  These must never be joined with the
    # preceding transaction, otherwise closing balance / grand total can become a
    # fake Tally voucher amount.
    prefixes = (
        "SOUTH INDIAN BANK LTD", "TRANSACTION DETAILS PAGE", "HTTP", "HTTPS",
        "MICR :", "IFSC :", "1ST FLOOR", "TO: PH:", "FAX:",
        "M/S.", "1 S C S C", "KAMRUP", "GUWAHATI", "ASSAM", "INDIA, PIN",
        "CKYC ID", "MODE OF OPR", "STATEMENT OF ACCOUNT", "DATE PARTICULARS",
        "PAGE TOTAL:", "GRAND TOTAL:", "EFF AVL AMT", "THIS IS AN AUTHENTICATED",
        "ACCOUNT HOLDERS", "END OF STATEMENT", "CLOSING BALANCE", "AVAILABLE BALANCE",
        "DATE:", "CUSTOMER ID:", "TYPE:", "A/C NO:", "CURRENCY CODE",
        "PAGE:", "CKYC", "EMAIL :",
    )
    if u.startswith(prefixes):
        return True
    if re.search(r"\b(GRAND TOTAL|PAGE TOTAL|EFF AVL AMT|AUTHENTICATED STATEMENT)\b", u):
        return True
    # Header date-time line such as: 05-05-2025 15:45:24 IFSC : SIBL0000473
    if re.match(r"^\d{2}[-/]\d{2}[-/]\d{4}\s+\d{2}:\d{2}:\d{2}\b", s):
        return True
    return False


def parse_south_indian_bank_text(text):
    """
    South Indian Bank Finacle PDF parser.

    Layout is text-based, not a proper PDF table:
      DATE PARTICULARS CHQ.NO. WITHDRAWALS DEPOSITS BALANCE
      02-04-25 Cash - GUWAHATI - NAME 1086569 10,000.00 6,55,319.82Cr

    The same line does not clearly preserve withdrawal/deposit columns after PDF
    text extraction. Therefore the reliable rule is balance movement:
      balance decreased => Withdrawal/Debit
      balance increased => Deposit/Credit
    """
    lines = [re.sub(r"\s+", " ", line).strip() for line in str(text or "").splitlines()]
    date_re = re.compile(r"^(?P<date>\d{2}[-/]\d{2}[-/]\d{2,4})\s+(?P<rest>.*)$")
    bal_re = re.compile(r"([\d,]+\.\d{2})\s*(Cr|Dr)\b", re.IGNORECASE)
    amt_re = re.compile(r"([\d,]+\.\d{2})")

    txns = []
    current = None
    prev_balance_signed = None

    def flush_current():
        nonlocal current, prev_balance_signed
        if not current:
            return
        joined = re.sub(r"\s+", " ", " ".join(current.get("parts", []))).strip()
        bal_matches = list(bal_re.finditer(joined))
        if not bal_matches:
            current = None
            return
        bal_m = bal_matches[-1]
        balance_signed = _sib_signed_balance(bal_m.group(0))
        if balance_signed is None:
            current = None
            return

        before_bal = joined[:bal_m.start()].strip()
        after_bal = joined[bal_m.end():].strip()
        amount_matches = list(amt_re.finditer(before_bal))
        if not amount_matches:
            current = None
            return
        amt_m = amount_matches[-1]
        amount = clean_amount(amt_m.group(1))

        narration = (before_bal[:amt_m.start()] + " " + before_bal[amt_m.end():] + " " + after_bal).strip()
        narration = re.sub(r"\s+", " ", narration).strip(" -|/") or "Bank Transaction"

        debit = credit = 0.0
        if prev_balance_signed is not None:
            delta = round(balance_signed - prev_balance_signed, 2)
            if abs(abs(delta) - round(amount, 2)) <= 2.0:
                if delta > 0:
                    credit = amount
                elif delta < 0:
                    debit = amount

        if debit == 0.0 and credit == 0.0:
            low = narration.lower()
            credit_hints = ["neft:", "nach_cr", "transfer:", "imps", "cash dep", "deposit"]
            debit_hints = ["charges", "chrgs", "charge", "int.coll", "int coll", "interest collected", "atm trn", "cwd", "pos trn", "loan recovery", "clearing:", "neftutr", "rtgs to"]
            if any(k in low for k in credit_hints) and not any(k in low for k in debit_hints):
                credit = amount
            else:
                debit = amount

        txns.append({
            "date": current.get("date"),
            "narration": narration,
            "voucher_type": detect_voucher_type(narration, debit, credit),
            "debit": round(float(debit or 0), 2),
            "credit": round(float(credit or 0), 2),
            "balance": round(abs(float(balance_signed or 0)), 2),
        })
        prev_balance_signed = balance_signed
        current = None

    for raw_line in lines:
        if not raw_line:
            continue

        if _is_sib_control_line(raw_line):
            # Do not let summary/footer/header lines become part of the previous
            # transaction.  Flush first so a real transaction just before a footer
            # (e.g. Int.Coll on 30-04-25) is captured with its own amount only.
            flush_current()
            continue

        m = date_re.match(raw_line)
        if m:
            dt = parse_date(m.group("date"))
            rest = (m.group("rest") or "").strip()
            # Opening balance line: 01-04-25 B/F 6,76,639.82Cr
            if re.match(r"^B/F\b", rest, re.IGNORECASE):
                bal_m = bal_re.search(rest)
                if bal_m:
                    prev_balance_signed = _sib_signed_balance(bal_m.group(0))
                current = None
                continue
            flush_current()
            if dt:
                current = {"date": dt, "parts": [rest]}
            continue

        if current is not None:
            current.setdefault("parts", []).append(raw_line)

    flush_current()
    return txns


# ── HDFC TEXT PARSER ─────────────────────────────────────────
def _extract_hdfc_opening_balance(text):
    m = re.search(
        r"Opening Balance\s+Dr Count\s+Cr Count\s+Debits\s+Credits\s+Closing Bal\s+([\d,]+\.\d{2})",
        str(text or ""),
        re.IGNORECASE,
    )
    if m:
        return clean_amount(m.group(1))
    return None


def parse_hdfc_text(text):
    """
    HDFC statement parser.

    HDFC PDFs visually have separate Withdrawal Amt. and Deposit Amt. columns.
    During PDF text/table extraction, blank cells often disappear, so the row may
    only show amount + closing balance. This caused deposits to be treated as
    payments. The safe rule is:
        closing balance increased by amount  => Receipt
        closing balance decreased by amount  => Payment
    """
    raw_text = str(text or "")
    lines = [re.sub(r"\s+", " ", line).strip() for line in raw_text.splitlines()]
    lines = [line for line in lines if line]

    tx_start_re = re.compile(r"^\s*(?P<date>\d{2}/\d{2}/\d{2,4})\s+(?P<rest>.+)$")
    val_amt_bal_re = re.compile(
        r"(?P<value>\d{2}/\d{2}/\d{2,4})\s+"
        r"(?P<amount>[\d,]+\.\d{2})\s+"
        r"(?P<balance>-?[\d,]+\.\d{2})(?:\s|$)"
    )

    control_prefixes = (
        "Page No", "Statement of account", "M/S.", "C/O", "FOOT HILL", "NURSERY",
        "GUWAHATI", "ASSAM", "JOINT HOLDERS", "Nomination", "Statement From",
        "Account Branch", "Address :", "City :", "State :", "Phone no", "OD Limit",
        "Email :", "Cust ID", "Account No", "A/C Open Date", "Account Status",
        "RTGS/NEFT", "Branch Code", "Account Type", "HDFC BANK LIMITED",
        "*Closing balance", "Contents of", "State account", "Registered Office",
        "Date Narration", "Withdrawal Amt", "Deposit Amt", "Closing Balance",
        "STATEMENT SUMMARY", "Opening Balance", "Generated On",
        "This is a computer generated", "not require signature",
        "JASWANTA ROAD", "this statement.", "Generated By",
        "Requesting Branch Code",
    )

    records = []
    current = None

    def flush_current():
        nonlocal current
        if not current:
            return

        joined = re.sub(r"\s+", " ", " ".join(current.get("parts", []))).strip()
        matches = list(val_amt_bal_re.finditer(joined))
        if not matches:
            current = None
            return

        m = matches[-1]
        value_dt = parse_date(m.group("value"))
        amount = clean_amount(m.group("amount"))
        balance = clean_amount(m.group("balance"))

        prefix = joined[:m.start()].strip()
        suffix = joined[m.end():].strip()

        ref = ""
        ref_m = re.search(r"^(?P<narr>.*?)(?:\s+)(?P<ref>[A-Z0-9][A-Z0-9/\-]{5,})$", prefix)
        if ref_m:
            narration = ref_m.group("narr").strip()
            ref = ref_m.group("ref").strip()
        else:
            narration = prefix

        if suffix:
            narration = (narration + " " + suffix).strip()

        narration = re.sub(r"\s+", " ", narration).strip(" -|") or "Bank Transaction"

        records.append({
            "date": current.get("date"),
            "value_date": value_dt,
            "narration": narration,
            "ref": ref,
            "amount": amount,
            "balance": balance,
        })
        current = None

    for line in lines:
        # Summary amount row: 30,361.74 301 741 7,218,214.78 ...
        if re.match(r"^[\d,]+\.\d{2}\s+\d+\s+\d+\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}\s+[\d,]+\.\d{2}$", line):
            continue
        if line.startswith(control_prefixes) or "HDFC Bank GSTIN" in line:
            continue

        m = tx_start_re.match(line)
        if m and parse_date(m.group("date")):
            flush_current()
            current = {
                "date": parse_date(m.group("date")),
                "parts": [m.group("rest").strip()],
            }
        else:
            if current:
                current["parts"].append(line)

    flush_current()

    txns = []
    prev_balance = _extract_hdfc_opening_balance(raw_text)

    for r in records:
        amount = float(r.get("amount", 0) or 0)
        balance = float(r.get("balance", 0) or 0)
        narration = str(r.get("narration") or "")
        narration_l = narration.lower()

        debit = credit = 0.0

        if prev_balance is not None:
            delta = round(balance - float(prev_balance), 2)
            if abs(delta - amount) <= 1.0:
                credit = amount
            elif abs(delta + amount) <= 1.0:
                debit = amount

        if debit == 0.0 and credit == 0.0:
            s = " " + narration_l
            credit_hints = [
                " neft cr", "neft cr-", "upi settlement", "rvsl", "refund",
                "cash deposit", "chq dep", "credit", "received", "deposit",
                "googleindiadigital", "google india digital", "accountmein fintech",
            ]
            debit_hints = [
                "chq paid", "ft - dr", "neft dr", "tpt-salery", "salary",
                "rental", "gst/bank", "paid", "debit", "charges", "withdrawal",
                "dr -", "i/w chq return",
            ]
            if any(k in s for k in credit_hints):
                credit = amount
            elif any(k in s for k in debit_hints):
                debit = amount
            elif re.search(r"\bcr\b", narration_l):
                credit = amount
            elif re.search(r"\bdr\b", narration_l):
                debit = amount
            else:
                debit = amount

        prev_balance = balance

        narr_out = narration
        if r.get("ref"):
            narr_out = f"{narr_out} | Ref: {r.get('ref')}"

        txns.append({
            "date": r.get("date"),
            "narration": narr_out,
            "voucher_type": detect_voucher_type(narr_out, debit, credit),
            "debit": round(float(debit or 0), 2),
            "credit": round(float(credit or 0), 2),
            "balance": round(float(balance or 0), 2),
        })

    return txns


# ── ICICI TEXT PARSER ─────────────────────────────────────────
def _looks_like_txn_date_token(token):
    return parse_date(str(token).strip()) is not None


def _icici_start_line(line):
    s = re.sub(r"\s+", " ", str(line or "").strip())
    if not s:
        return False
    # Common ICICI formats:
    # 01/04/2026 ...
    # 1 01/04/2026 01/04/2026 ...
    # 01-Apr-2026 ...
    date_pat = r"(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\.\d{1,2}\.\d{2,4}|\d{1,2}[-/ ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*[-/ ]\d{2,4})"
    return bool(re.match(rf"^(?:\d+\s+)?{date_pat}\b", s, re.IGNORECASE))


def _icici_control_line(line):
    s = re.sub(r"\s+", " ", str(line or "").strip())
    u = s.upper()
    if not s:
        return True
    if _icici_start_line(s):
        return False
    prefixes = (
        "ICICI BANK", "STATEMENT", "ACCOUNT STATEMENT", "STATEMENT OF", "CUSTOMER ID",
        "ACCOUNT NO", "A/C NO", "ACCOUNT NUMBER", "BRANCH", "ADDRESS", "PHONE",
        "EMAIL", "IFSC", "MICR", "NOMINATION", "NOMINEE", "PAGE", "DATE PARTICULARS",
        "DATE MODE", "TXN DATE", "VALUE DATE", "TRANSACTION DATE", "TRANSACTION REMARKS",
        "PARTICULARS CHQ", "WITHDRAWAL", "WITHDRAWALS", "DEPOSIT", "DEPOSITS",
        "BALANCE", "OPENING BALANCE", "CLOSING BALANCE", "TOTAL", "SUMMARY",
        "THIS IS A COMPUTER", "REGISTERED OFFICE", "END OF STATEMENT", "LEGENDS",
        "FOR ANY QUERIES", "IMPORTANT INFORMATION", "PLEASE NOTE", "GENERATED",
    )
    if u.startswith(prefixes):
        return True
    if "WITHDRAWAL" in u and "DEPOSIT" in u and "BALANCE" in u:
        return True
    return False


def _icici_amount_tokens_with_span(text_line):
    # Amounts in ICICI statements normally have decimals. Handles optional CR/DR suffix.
    token_re = re.compile(r"(?<![A-Za-z0-9])(?:INR\s*)?(-?\d{1,3}(?:,\d{2,3})*(?:\.\d{2})|-?\d+\.\d{2})(?:\s*(Cr|Dr))?(?![A-Za-z0-9])", re.IGNORECASE)
    out = []
    for m in token_re.finditer(str(text_line or "")):
        raw = m.group(1)
        tag = (m.group(2) or "").upper()
        val = clean_amount(raw)
        out.append({"value": val, "raw": raw, "tag": tag, "start": m.start(), "end": m.end()})
    return out


def _icici_extract_dates_and_rest(row_text):
    s = re.sub(r"\s+", " ", str(row_text or "").strip())
    # Remove leading serial number where present
    s = re.sub(r"^\d+\s+(?=\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}[-/ ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec))", "", s, flags=re.IGNORECASE)
    date_pat = r"(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\.\d{1,2}\.\d{2,4}|\d{1,2}[-/ ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*[-/ ]\d{2,4})"
    dates = []
    for _ in range(2):
        m = re.match(rf"^({date_pat})\b\s*", s, re.IGNORECASE)
        if not m:
            break
        dates.append(m.group(1))
        s = s[m.end():].strip()
    dt = parse_date(dates[0]) if dates else None
    return dt, s


def _icici_side_from_narration(narration):
    s = " " + re.sub(r"\s+", " ", str(narration or "").lower()) + " "
    credit_hints = (
        " deposit", "cash dep", "by cash", "by clg", "clearing", "credit", " cr ",
        "received", "receipt", "salary", "neft from", "rtgs from", "imps from",
        "upi cr", "refund", "interest paid", "inward", "ecs cr", "cms",
    )
    debit_hints = (
        "withdraw", "debit", " dr ", "charges", "charge", "chgs", "paid", "payment",
        "upi/", "upi-", "imps/", "neft/", "rtgs/", "atm", "pos/", "to ",
        "transfer to", "fund trf", "clg chq", "cheque paid", "ach debit", "ecs dr",
    )
    if any(k in s for k in credit_hints):
        return "credit"
    if any(k in s for k in debit_hints):
        return "debit"
    return "debit"


def _icici_find_opening_balance(text):
    raw = str(text or "")
    patterns = [
        r"Opening\s+Balance[^\d-]*(-?\d{1,3}(?:,\d{2,3})*(?:\.\d{2})|-?\d+\.\d{2})\s*(Cr|Dr)?",
        r"Balance\s+B/F[^\d-]*(-?\d{1,3}(?:,\d{2,3})*(?:\.\d{2})|-?\d+\.\d{2})\s*(Cr|Dr)?",
    ]
    for pat in patterns:
        m = re.search(pat, raw, re.IGNORECASE)
        if m:
            val = clean_amount(m.group(1))
            tag = (m.group(2) or "").lower()
            if tag == "dr":
                val = -val
            return val
    return None


def _icici_balance_value(token):
    val = float(token.get("value", 0) or 0)
    if str(token.get("tag", "")).upper() == "DR":
        return -val
    return val


def parse_icici_text(text):
    raw_text = str(text or "")
    lines = [re.sub(r"\s+", " ", line).strip() for line in raw_text.splitlines()]
    lines = [line for line in lines if line and not _icici_control_line(line)]

    rows = []
    current = None
    for line in lines:
        if _icici_start_line(line):
            if current:
                rows.append(" ".join(current))
            current = [line]
        elif current:
            # Continuation narration line
            current.append(line)
    if current:
        rows.append(" ".join(current))

    txns = []
    prev_balance = _icici_find_opening_balance(raw_text)

    for row in rows:
        dt, rest = _icici_extract_dates_and_rest(row)
        if not dt or not rest:
            continue

        tokens = _icici_amount_tokens_with_span(rest)
        if len(tokens) < 2:
            continue

        # The last amount is normally the running balance. Amounts before that are withdrawal/deposit.
        bal_token = tokens[-1]
        balance = _icici_balance_value(bal_token)
        amount_tokens = tokens[:-1]

        # Narration is text before the first transaction amount. Remove stray cheque/reference tail only if needed.
        narration = rest[:amount_tokens[0]["start"]].strip()
        if not narration:
            narration = re.sub(r"\s+", " ", rest[:bal_token["start"]]).strip()
        narration = re.sub(r"\s+", " ", narration).strip(" -|/") or "Bank Transaction"

        debit = credit = 0.0

        # If ICICI text preserved both columns, convention is Withdrawal then Deposit before Balance.
        nonzero_amounts = [t for t in amount_tokens if abs(float(t.get("value", 0) or 0)) > 0.0001]
        if len(amount_tokens) >= 2:
            withdrawal = float(amount_tokens[-2].get("value", 0) or 0)
            deposit = float(amount_tokens[-1].get("value", 0) or 0)
            if withdrawal > 0 and deposit == 0:
                debit = withdrawal
            elif deposit > 0 and withdrawal == 0:
                credit = deposit
            elif len(nonzero_amounts) == 1:
                amount = float(nonzero_amounts[0].get("value", 0) or 0)
                # Use balance movement if opening/previous balance available.
                if prev_balance is not None:
                    delta = round(balance - float(prev_balance), 2)
                    if abs(delta - amount) <= 1.0:
                        credit = amount
                    elif abs(delta + amount) <= 1.0:
                        debit = amount
                if debit == 0.0 and credit == 0.0:
                    if _icici_side_from_narration(narration) == "credit":
                        credit = amount
                    else:
                        debit = amount
            else:
                # Rare case: both columns have values. Keep both rather than losing data.
                debit = withdrawal
                credit = deposit
        else:
            amount = float(amount_tokens[0].get("value", 0) or 0)
            tag = str(amount_tokens[0].get("tag", "")).upper()
            if tag == "CR":
                credit = amount
            elif tag == "DR":
                debit = amount
            elif prev_balance is not None:
                delta = round(balance - float(prev_balance), 2)
                if abs(delta - amount) <= 1.0:
                    credit = amount
                elif abs(delta + amount) <= 1.0:
                    debit = amount
            if debit == 0.0 and credit == 0.0:
                if _icici_side_from_narration(narration) == "credit":
                    credit = amount
                else:
                    debit = amount

        prev_balance = balance
        if debit == 0.0 and credit == 0.0:
            continue

        txns.append({
            "date": dt,
            "narration": narration,
            "voucher_type": detect_voucher_type(narration, debit, credit),
            "debit": round(float(debit or 0), 2),
            "credit": round(float(credit or 0), 2),
            "balance": round(abs(float(balance or 0)), 2),
        })

    return txns


def parse_icici_table_rows(tables):
    """Fallback for ICICI tables when text rows are split into proper columns."""
    txns = []
    header_map = {}
    for row in tables or []:
        if not row:
            continue
        cells = [re.sub(r"\s+", " ", str(c or "")).strip() for c in row]
        lower = [c.lower() for c in cells]

        # Detect header anywhere in the statement.
        if any("withdraw" in c or "debit" in c for c in lower) and any("deposit" in c or "credit" in c for c in lower):
            header_map = {}
            for i, c in enumerate(lower):
                if "date" in c and "value" not in c and "date" not in header_map:
                    header_map["date"] = i
                if "value date" in c:
                    header_map.setdefault("value_date", i)
                if any(k in c for k in ["particular", "remark", "description", "narration"]):
                    header_map["narration"] = i
                if any(k in c for k in ["withdraw", "debit"]):
                    header_map["debit"] = i
                if any(k in c for k in ["deposit", "credit"]):
                    header_map["credit"] = i
                if "balance" in c:
                    header_map["balance"] = i
            continue

        if not header_map:
            continue

        date_idx = header_map.get("date", header_map.get("value_date", 0))
        dt = parse_date(cells[date_idx]) if date_idx < len(cells) else None
        if not dt:
            # Try any date cell in this row.
            for c in cells[:4]:
                dt = parse_date(c)
                if dt:
                    break
        if not dt:
            continue

        narr_idx = header_map.get("narration", 2)
        narration = cells[narr_idx] if narr_idx < len(cells) else "Bank Transaction"
        narration = re.sub(r"\s+", " ", narration).strip() or "Bank Transaction"
        debit_idx = header_map.get("debit", -1)
        credit_idx = header_map.get("credit", -1)
        balance_idx = header_map.get("balance", -1)
        debit = _parse_amount(cells[debit_idx]) if 0 <= debit_idx < len(cells) else 0.0
        credit = _parse_amount(cells[credit_idx]) if 0 <= credit_idx < len(cells) else 0.0
        balance = _parse_amount(cells[balance_idx]) if 0 <= balance_idx < len(cells) else 0.0
        if debit == 0.0 and credit == 0.0:
            continue
        txns.append({
            "date": dt,
            "narration": narration,
            "voucher_type": detect_voucher_type(narration, debit, credit),
            "debit": round(float(debit or 0), 2),
            "credit": round(float(credit or 0), 2),
            "balance": round(float(balance or 0), 2),
        })
    return txns


# ── PARSE FROM TEXT ──────────────────────────────────────────
def parse_from_text(text, bank=None):
    bank_name = (bank or detect_bank(text) or "").upper()
    if bank_name in ["SOUTH INDIAN BANK", "SIB"]:
        return parse_south_indian_bank_text(text)
    if bank_name == "INDIAN BANK":
        return parse_indian_bank_text(text)
    if bank_name == "HDFC":
        return parse_hdfc_text(text)
    if bank_name == "ICICI":
        return parse_icici_text(text)

    txns    = []
    date_re = re.compile(r"\b(\d{2}[\/\-]\d{2}[\/\-]\d{2,4})\b")
    amt_re  = re.compile(r"[\d,]+\.\d{2}")
    for line in text.split("\n"):
        dm = date_re.search(line)
        if not dm:
            continue
        dt = parse_date(dm.group(1))
        if not dt:
            continue
        amounts = [clean_amount(m) for m in amt_re.findall(line)]
        if not amounts:
            continue
        narration = re.sub(r"\s+", " ",
                           amt_re.sub("", line.replace(dm.group(1), "")).strip()
                           ) or "Bank Transaction"
        is_cr  = bool(re.search(r"\bcr\b|credit|deposit|received|salary", line.lower()))
        debit  = credit = 0.0
        balance = amounts[-1]
        if len(amounts) >= 2:
            credit, debit = (amounts[0], 0) if is_cr else (0, amounts[0])
        else:
            credit, debit = (amounts[0], 0) if is_cr else (0, amounts[0])

        txns.append({
            "date":         dt,
            "narration":    narration,
            "voucher_type": detect_voucher_type(narration, debit, credit),
            "debit":        debit,
            "credit":       credit,
            "balance":      balance,
        })
    return txns


# ── DUPLICATE DETECTION ──────────────────────────────────────
def find_duplicates(txns):
    seen, dupes = {}, []
    for i, t in enumerate(txns):
        key = (
            t["date"].strftime("%Y%m%d") if t["date"] else "",
            round(t["debit"],   2),
            round(t["credit"],  2),
            round(t["balance"], 2) if t.get("balance") else "",
            t["narration"][:40],
        )
        if key in seen:
            dupes.append((seen[key], i))
        else:
            seen[key] = i
    return dupes


# ── DATE FILTER ──────────────────────────────────────────────
def filter_by_date(txns, start_date=None, end_date=None):
    result = []
    for t in txns:
        if not t["date"]:
            continue
        if start_date and t["date"] < start_date:
            continue
        if end_date and t["date"] > end_date:
            continue
        result.append(t)
    return result


# ── TALLY XML EXPORT ─────────────────────────────────────────
def export_xml(txns, bank_name, output_path, tally_ledger="", tally_company=""):
    """
    Tally Prime compatible XML format.
    Import: Gateway of Tally > Import > Data > Vouchers
    tally_ledger: Exact Tally mein jo ledger name hai
    tally_company: Safe Push ke liye loaded Tally company name. Manual XML export
                   ke liye blank allowed hai; blank SVCurrentCompany tag emit nahi hoga.
    """
    def _x(v):
        return str(v or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    bank_ledger = tally_ledger if tally_ledger else f"{bank_name} Bank A/c"
    bank_ledger_xml = _x(bank_ledger)
    lines = []
    lines.append('<?xml version="1.0" encoding="UTF-8"?>')
    lines.append('<ENVELOPE>')
    lines.append('  <HEADER>')
    lines.append('    <TALLYREQUEST>Import Data</TALLYREQUEST>')
    lines.append('  </HEADER>')
    lines.append('  <BODY>')
    lines.append('    <IMPORTDATA>')
    lines.append('      <REQUESTDESC>')
    lines.append('        <REPORTNAME>Vouchers</REPORTNAME>')
    # Blank SVCURRENTCOMPANY direct HTTP push me Tally error deta hai:
    # Could not set 'SVCurrentCompany' to ''. Blank ho to tag omit karo.
    tally_company_xml = _x(tally_company).strip()
    if tally_company_xml:
        lines.append('        <STATICVARIABLES>')
        lines.append(f'          <SVCURRENTCOMPANY>{tally_company_xml}</SVCURRENTCOMPANY>')
        lines.append('        </STATICVARIABLES>')
    lines.append('      </REQUESTDESC>')
    lines.append('      <REQUESTDATA>')

    for idx, t in enumerate(txns, 1):
        # Stable voucher sequence: important for Safe Push (one-by-one) and repeat XML export.
        # convert() stores _bank_import_seq for every extracted bank row.
        try:
            seq_no = int(t.get("_bank_import_seq") or idx)
        except Exception:
            seq_no = idx
        amount = t["debit"] if t["debit"] > 0 else t["credit"]
        if amount == 0:
            continue

        vtype      = t["voucher_type"]
        is_debit   = t["debit"] > 0
        subtype    = t.get("contra_subtype", "")
        tally_date = format_tally_date(t["date"])
        narr = _x(t.get("narration", ""))

        if vtype == "Contra":
            # Cash Deposit: Cash aaya bank mein → Dr Bank, Cr Cash
            # Cash Withdrawal: Cash nikla bank se → Dr Cash, Cr Bank
            amt = round(amount, 2)
            if subtype == "cash_deposit":
                # Bank Dr (+), Cash Cr (-)
                bank_amt = -amt;  bank_pos = "Yes"
                cash_amt =  amt;  cash_pos = "No"
                cash_ledger = _x("Cash")
            else:
                # Bank Cr (-), Cash Dr (+)
                bank_amt =  amt;  bank_pos = "No"
                cash_amt = -amt;  cash_pos = "Yes"
                cash_ledger = _x("Cash")
            lines.append(f'        <TALLYMESSAGE xmlns:UDF="TallyUDF">')
            lines.append(f'          <VOUCHER VCHTYPE="Contra" ACTION="Create" OBJVIEW="Accounting Voucher View">')
            lines.append(f'            <DATE>{tally_date}</DATE>')
            lines.append(f'            <GUID>BNK-{seq_no:06d}-{tally_date}</GUID>')
            lines.append(f'            <NARRATION>{narr}</NARRATION>')
            lines.append(f'            <VOUCHERTYPENAME>Contra</VOUCHERTYPENAME>')
            lines.append(f'            <VOUCHERNUMBER>BNK/{seq_no:04d}</VOUCHERNUMBER>')
            lines.append(f'            <REFERENCE>{narr}</REFERENCE>')
            lines.append(f'            <ISINVOICE>No</ISINVOICE>')
            lines.append(f'            <HASCASHFLOW>Yes</HASCASHFLOW>')
            lines.append(f'            <PERSISTEDVIEW>Accounting Voucher View</PERSISTEDVIEW>')
            lines.append(f'            <ALLLEDGERENTRIES.LIST>')
            lines.append(f'              <LEDGERNAME>{bank_ledger_xml}</LEDGERNAME>')
            lines.append(f'              <ISDEEMEDPOSITIVE>{bank_pos}</ISDEEMEDPOSITIVE>')
            lines.append(f'              <ISPARTYLEDGER>No</ISPARTYLEDGER>')
            lines.append(f'              <AMOUNT>{bank_amt}</AMOUNT>')
            lines.append(f'            </ALLLEDGERENTRIES.LIST>')
            lines.append(f'            <ALLLEDGERENTRIES.LIST>')
            lines.append(f'              <LEDGERNAME>{cash_ledger}</LEDGERNAME>')
            lines.append(f'              <ISDEEMEDPOSITIVE>{cash_pos}</ISDEEMEDPOSITIVE>')
            lines.append(f'              <ISPARTYLEDGER>No</ISPARTYLEDGER>')
            lines.append(f'              <AMOUNT>{cash_amt}</AMOUNT>')
            lines.append(f'            </ALLLEDGERENTRIES.LIST>')
            lines.append(f'          </VOUCHER>')
            lines.append(f'        </TALLYMESSAGE>')
            continue

        # Payment / Receipt
        # Ledger Matching Tool final confirmed ledger ko yahan use karega.
        # Confirm nahi hua to safe fallback = Suspense Account.
        opp_ledger = (t.get("opposite_ledger") or t.get("matched_ledger") or "Suspense Account").strip()
        if not t.get("ledger_match_confirmed") and opp_ledger != "Suspense Account":
            opp_ledger = "Suspense Account"
        opp_ledger_xml = _x(opp_ledger)
        if is_debit:
            bank_amt =  round(amount, 2); opp_amt = -round(amount, 2)
            bank_pos = "No";  opp_pos = "Yes"
        else:
            bank_amt = -round(amount, 2); opp_amt =  round(amount, 2)
            bank_pos = "Yes"; opp_pos = "No"

        lines.append(f'        <TALLYMESSAGE xmlns:UDF="TallyUDF">')
        lines.append(f'          <VOUCHER VCHTYPE="{vtype}" ACTION="Create" OBJVIEW="Accounting Voucher View">')
        lines.append(f'            <DATE>{tally_date}</DATE>')
        lines.append(f'            <GUID>BNK-{seq_no:06d}-{tally_date}</GUID>')
        lines.append(f'            <NARRATION>{narr}</NARRATION>')
        lines.append(f'            <VOUCHERTYPENAME>{vtype}</VOUCHERTYPENAME>')
        lines.append(f'            <VOUCHERNUMBER>BNK/{seq_no:04d}</VOUCHERNUMBER>')
        lines.append(f'            <REFERENCE>{narr}</REFERENCE>')
        lines.append(f'            <ISINVOICE>No</ISINVOICE>')
        lines.append(f'            <PERSISTEDVIEW>Accounting Voucher View</PERSISTEDVIEW>')
        lines.append(f'            <ALLLEDGERENTRIES.LIST>')
        lines.append(f'              <LEDGERNAME>{bank_ledger_xml}</LEDGERNAME>')
        lines.append(f'              <ISDEEMEDPOSITIVE>{bank_pos}</ISDEEMEDPOSITIVE>')
        lines.append(f'              <ISPARTYLEDGER>No</ISPARTYLEDGER>')
        lines.append(f'              <AMOUNT>{bank_amt}</AMOUNT>')
        lines.append(f'            </ALLLEDGERENTRIES.LIST>')
        lines.append(f'            <ALLLEDGERENTRIES.LIST>')
        lines.append(f'              <LEDGERNAME>{opp_ledger_xml}</LEDGERNAME>')
        lines.append(f'              <ISDEEMEDPOSITIVE>{opp_pos}</ISDEEMEDPOSITIVE>')
        lines.append(f'              <ISPARTYLEDGER>No</ISPARTYLEDGER>')
        lines.append(f'              <AMOUNT>{opp_amt}</AMOUNT>')
        lines.append(f'            </ALLLEDGERENTRIES.LIST>')
        lines.append(f'          </VOUCHER>')
        lines.append(f'        </TALLYMESSAGE>')

    lines.append('      </REQUESTDATA>')
    lines.append('    </IMPORTDATA>')
    lines.append('  </BODY>')
    lines.append('</ENVELOPE>')

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ── CSV EXPORT ───────────────────────────────────────────────
def export_csv(txns, output_path):
    if pd is None:
        raise RuntimeError("CSV export dependency is not installed on the Phase 3 preview server.")
    rows = [{
        "Date":        t["date"].strftime("%d/%m/%Y") if t["date"] else "",
        "Narration":   t["narration"],
        "VoucherType": t["voucher_type"],
        "Debit":       t["debit"]   if t["debit"]   > 0 else "",
        "Credit":      t["credit"]  if t["credit"]  > 0 else "",
        "Balance":     t["balance"] if t["balance"] > 0 else "",
    } for t in txns]
    pd.DataFrame(rows).to_csv(output_path, index=False, encoding="utf-8-sig")


# ── EXCEL EXPORT ─────────────────────────────────────────────
def export_excel(txns, bank_name, output_path):
    if openpyxl is None:
        raise RuntimeError("Excel export dependency is not installed on the Phase 3 preview server.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    hdr_fill    = PatternFill("solid", start_color="1a2035")
    hdr_font    = Font(bold=True, color="63b3ed", size=11)
    hdr_align   = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="2d3748"),
        top   =Side(style="thin", color="2d3748"),
    )
    debit_font  = Font(color="E53E3E", size=10)
    credit_font = Font(color="38A169", size=10)
    contra_font = Font(color="D69E2E", size=10)
    normal_font = Font(size=10)
    alt_fill    = PatternFill("solid", start_color="f7f9fc")

    ws.merge_cells("A1:G1")
    ws["A1"]           = f"BankSync Pro — {bank_name} Bank Statement"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", start_color="0d1526")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Date", "Narration", "Voucher Type", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=2, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = thin_border
    ws.row_dimensions[2].height = 22

    total_debit = total_credit = 0
    payment_count = receipt_count = 0

    for row_idx, t in enumerate(txns, 3):
        fill  = alt_fill if row_idx % 2 == 0 else PatternFill()
        vtype = t["voucher_type"]
        if vtype == "Payment":
            row_font = debit_font;  payment_count += 1
        else:
            row_font = credit_font; receipt_count += 1

        ws.cell(row=row_idx, column=1,
                value=t["date"].strftime("%d/%m/%Y") if t["date"] else "").font = normal_font
        ws.cell(row=row_idx, column=2, value=t["narration"]).font = normal_font
        ws.cell(row=row_idx, column=3, value=vtype).font          = row_font

        d = ws.cell(row=row_idx, column=4, value=t["debit"]   if t["debit"]   > 0 else "")
        d.font = debit_font;  d.number_format = '#,##0.00'
        c = ws.cell(row=row_idx, column=5, value=t["credit"]  if t["credit"]  > 0 else "")
        c.font = credit_font; c.number_format = '#,##0.00'
        bal_val = t.get("balance", "")
        b = ws.cell(row=row_idx, column=6, value=bal_val if bal_val not in ("", None) else "")
        b.font = normal_font; b.number_format = '#,##0.00'

        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = fill

        total_debit  += t["debit"]
        total_credit += t["credit"]

    last = len(txns) + 3
    ws.cell(row=last, column=3, value="TOTAL").font = Font(bold=True, size=10)
    td = ws.cell(row=last, column=4, value=total_debit)
    td.font = Font(bold=True, color="E53E3E"); td.number_format = '#,##0.00'
    tc = ws.cell(row=last, column=5, value=total_credit)
    tc.font = Font(bold=True, color="38A169"); tc.number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "BankSync Pro — Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    for r, (k, v) in enumerate([
        ("Bank",               f"{bank_name} Bank"),
        ("Total Transactions", len(txns)),
        ("Payment",            payment_count),
        ("Receipt",            receipt_count),
        ("Total Debit",        total_debit),
        ("Total Credit",       total_credit),
        ("Net",                total_credit - total_debit),
        ("Date From",          txns[0]["date"].strftime("%d/%m/%Y")  if txns else ""),
        ("Date To",            txns[-1]["date"].strftime("%d/%m/%Y") if txns else ""),
    ], 3):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 22
    wb.save(output_path)


# ── PROFILES ─────────────────────────────────────────────────
def load_profiles():
    if os.path.exists(PROFILES_FILE):
        with open(PROFILES_FILE) as f:
            return json.load(f)
    return {}


def save_profile(name, data):
    profiles = load_profiles()
    profiles[name] = data
    with open(PROFILES_FILE, "w") as f:
        json.dump(profiles, f, indent=2)


# ── MAIN CONVERT ─────────────────────────────────────────────
def convert(pdf_path, bank=None, password="", start_date=None, end_date=None,
            skip_duplicates=False, export_formats=None, progress_cb=None,
            tally_ledger=""):

    if export_formats is None:
        export_formats = ["xml"]

    if password:
        pdf_path = unlock_pdf(pdf_path, password)

    # Fast text path: HDFC and ICICI statements often parse better from raw text
    # because PDF table extraction can collapse Withdrawal/Deposit columns.
    selected_bank = (str(bank or "").strip().upper())
    ledger_hint = (str(tally_ledger or "").strip().upper())
    if selected_bank in ["", "AUTO"] and "ICICI" in ledger_hint:
        selected_bank = "ICICI"
    if selected_bank in ["", "AUTO"] and "HDFC" in ledger_hint:
        selected_bank = "HDFC"
    if selected_bank in ["", "AUTO"] and "IDBI" in ledger_hint:
        selected_bank = "IDBI"
    if selected_bank in ["", "AUTO"] and "UCO" in ledger_hint:
        selected_bank = "UCO"
    if selected_bank in ["", "AUTO"] and ("SOUTH INDIAN" in ledger_hint or "SIB" in ledger_hint):
        selected_bank = "SOUTH INDIAN BANK"

    fast_text = extract_pdf_text_fast(pdf_path, progress_cb=None, password=password)
    fast_bank = detect_bank(fast_text) if fast_text else ""

    if selected_bank == "HDFC" or ((selected_bank in ["", "AUTO"]) and fast_bank == "HDFC"):
        text, tables = fast_text, []
        bank = "HDFC"
    elif selected_bank == "ICICI" or ((selected_bank in ["", "AUTO"]) and fast_bank == "ICICI"):
        # Keep tables too as a fallback, but prefer raw text for transaction rows.
        text, tables = extract_pdf(pdf_path, progress_cb, password=password)
        if fast_text and len(fast_text) > len(text or "") * 0.75:
            text = fast_text
        bank = "ICICI"
    elif selected_bank == "IDBI" or ((selected_bank in ["", "AUTO"]) and fast_bank == "IDBI"):
        # IDBI GoMobile PDF is a proper table statement. Use pdfplumber tables.
        text, tables = extract_pdf(pdf_path, progress_cb, password=password)
        bank = "IDBI"
    elif selected_bank == "UCO" or ((selected_bank in ["", "AUTO"]) and fast_bank == "UCO"):
        # UCO statements use a fixed 6-column table without normal headers.
        text, tables = extract_pdf(pdf_path, progress_cb, password=password)
        bank = "UCO"
    elif selected_bank in ["SOUTH INDIAN BANK", "SIB"] or ((selected_bank in ["", "AUTO"]) and fast_bank in ["SOUTH INDIAN BANK", "SIB"]):
        # South Indian Bank PDF is text-layout and pdfplumber keeps transaction
        # line breaks better than pypdf for this Finacle report.
        text, tables = extract_pdf(pdf_path, progress_cb, password=password)
        if not text and fast_text:
            text = fast_text
        bank = "SOUTH INDIAN BANK"
    else:
        text, tables = extract_pdf(pdf_path, progress_cb, password=password)
        if not text and fast_text:
            text = fast_text
        if not bank or bank == "AUTO":
            bank = detect_bank(text)

    txns = []
    bank_upper = str(bank or "").upper()
    if bank_upper == "HDFC":
        txns = parse_hdfc_text(text)
    elif bank_upper == "ICICI":
        # ICICI detailed statements are table-based. Prefer table extraction because
        # raw text extraction can merge the header/date area with all rows and return
        # 0/1 incorrect transactions. Fallback to text parser only if tables fail.
        txns = parse_icici_table_rows(tables) if tables else []
        if not txns:
            txns = parse_icici_text(text)
    elif bank_upper == "UCO":
        txns = parse_uco_table(tables) if tables else []
    elif bank_upper in ["SOUTH INDIAN BANK", "SIB"]:
        txns = parse_south_indian_bank_text(text)
    elif bank_upper in ["IDBI", "AXIS"]:
        txns = parse_from_table(tables) if tables else []
    elif bank_upper != "INDIAN BANK":
        txns = parse_from_table(tables) if tables else []
    if not txns:
        txns = parse_from_text(text, bank=bank)

    if not txns:
        if not str(text or "").strip():
            raise RuntimeError("No readable text was found in this PDF. It looks scanned/image-based, so convert it to searchable/text PDF first or upload an Excel statement.")
        raise RuntimeError("No transactions could be extracted from this PDF. Please select the correct bank/ledger. If this still fails, the PDF layout may be image/scanned or non-standard.")

    if start_date or end_date:
        txns = filter_by_date(txns, start_date, end_date)

    txns.sort(key=lambda x: x["date"] or datetime.min)

    dupes = find_duplicates(txns)
    # IMPORTANT FIX:
    # Bank statements can contain genuine repeated transactions with the same
    # date, amount and narration (for example two separate ₹59 charges).
    # Earlier the app removed the later matching row when skip_duplicates=True,
    # so Tally received only one voucher.  We now keep every extracted bank row
    # by default; duplicate information is only reported for review.
    if skip_duplicates and dupes:
        dupe_indices = {j for _, j in dupes}
        txns = [t for i, t in enumerate(txns) if i not in dupe_indices]

    # Give every voucher a visible internal row sequence for safer matching/patching.
    for _seq, _txn in enumerate(txns, 1):
        _txn.setdefault("_bank_import_seq", _seq)

    # ── Contra detection PEHLE karo, phir XML generate ──────────
    import re as _re
    _CONTRA_KW = [
        "cash deposit", "cash deposited", "cash dep", "cash dep self", "cash deposit self",
        "cash depositself", "by cash", "cash -", "cash /", "cdm deposit", "cdm cash", "atm dep", "atm deposit",
        "cash withdrawal", "cash withdrawn", "cash withd", "cash wdl",
        "cash withdraw self", "cash withdrawal self", "cash withdrawal by self",
        "cash wd", "cash w/d", "atm withdrawal", "atm cash withdrawal",
        "atm wd", "atm wdl", "cash at atm", "cash withdrawl",
    ]
    _NON_CONTRA_KW = [
        "cash dep chg", "cash deposit chg", "cash handling", "inter city", "intercity",
        "charges", "charge", " chg ", " chgs ", "+gst", "fee", "fees", "commission",
        "interest", "gst on", "imps", "upi/cr", "upi/dr", "neft", "rtgs",
        "by transfer", "transfer from", "transfer to", "cheque deposit",
        "inb ", "inbct", "sbiy", "salary",
    ]
    for t in txns:
        raw = " ".join([str(t.get("narration") or ""), str(t.get("description") or ""),
                        str(t.get("particulars") or ""), str(t.get("remarks") or "")])
        narr = _re.sub(r"\s+", " ", _re.sub(r"[-_]+", " ", raw)).lower().strip()
        dr = float(t.get("debit", 0) or 0)
        cr = float(t.get("credit", 0) or 0)
        if any(kw in narr for kw in _NON_CONTRA_KW):
            continue
        is_contra = any(kw in narr for kw in _CONTRA_KW)
        if not is_contra:
            has_cash   = any(w in narr for w in ("cash", "atm", "cdm"))
            has_action = any(w in narr for w in ("deposit", "withdraw", "withdrawal", "withdrawl", "wdl", "self"))
            is_contra  = has_cash and has_action
        # South Indian Bank commonly prints cash withdrawals as only
        # "Cash - BRANCH - NAME". There is no WDL/SELF keyword, so treat
        # debit-side narration starting with Cash as cash withdrawal/Contra.
        if not is_contra and dr > 0 and _re.match(r"^cash\b", narr):
            is_contra = True
        if is_contra and (dr > 0 or cr > 0):
            t["_original_voucher_type"] = t.get("voucher_type", "")
            t["voucher_type"]  = "Contra"
            t["contra_subtype"] = "cash_deposit" if cr > 0 else "cash_withdrawal"
    # ─────────────────────────────────────────────────────────────

    base    = os.path.splitext(pdf_path)[0].replace("_unlocked", "")
    outputs = {}
    if "xml"   in export_formats:
        p = base + "_tally.xml";      export_xml(txns, bank, p, tally_ledger);   outputs["xml"]   = p
    if "csv"   in export_formats:
        p = base + "_statement.csv";  export_csv(txns, p);         outputs["csv"]   = p
    if "excel" in export_formats:
        p = base + "_statement.xlsx"; export_excel(txns, bank, p); outputs["excel"] = p

    return {
        "bank":          bank,
        "transactions":  txns,
        "duplicates":    dupes,
        "outputs":       outputs,
        "total":         len(txns),
        "total_debit":   sum(t["debit"]  for t in txns),
        "total_credit":  sum(t["credit"] for t in txns),
        "payment_count": sum(1 for t in txns if t["voucher_type"] == "Payment"),
        "receipt_count": sum(1 for t in txns if t["voucher_type"] == "Receipt"),
    }


# ═══════════════════════════════════════════════════════════════
#  GUI
# ═══════════════════════════════════════════════════════════════
